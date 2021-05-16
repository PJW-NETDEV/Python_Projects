###### Program to read Excel data in an array and upload somewhere using API
###### Need to provide path to .xlsx file
###### Usage: python3 ReadIPAM.py ~/Documents/IPAM.xlsx

import sys
import pprint
import os.path
from openpyxl import load_workbook
from array import *

# Initialise Variables (Not really required)

# Get Passed In variables From sys.argv and then assiged passed file to variable
my_arg_list = sys.argv
SourceFilePath = my_arg_list[1]
MyIPs =[]
MySheets =[]
LoopCounter = int(0)
SelectedSheet = str()
PreviewAmount = 10
KeyColumn = str()
StartRow = int()
EndRow = int()
NumRows = int()
EndRowNum = int()

# Create array for interating column references
letters = [
           "a","b","c","d","e","f","g","h","i","k",
           "l","m","n","o","p","q","r","s","t","u",
           "v","w","x","y","z"]

#Setup pp print
pp = pprint.PrettyPrinter(indent=4)

#Check if file Exists, import if OK.
if os.path.isfile(SourceFilePath):
    print("File Check: " + SourceFilePath + "   OK!\n")
    myworkbook = load_workbook(SourceFilePath, data_only=True)
else:
    print ("File not exist")
    exit()


# Assign References to Sheet Names in 2D array - for the menu
# Loop Counter used to provide an ID number for each sheet name
for ExcelSheetName in myworkbook.sheetnames:
    LoopCounter += 1
    MySheets.append([LoopCounter,ExcelSheetName])

#Display a menu to help select a sheet name - adding a space for numbers <10 to align sheet names
print("Sheets Names:")
for x in range(len(MySheets)):
    if x < 9:
        Padding_Char = " "
    else:
        Padding_Char = ""
    print(str(MySheets[x][0]) + Padding_Char + " -- " + str(MySheets[x][1]))


# Ask user for selected sheet ID, validation to check entry is a number and within range
while True:      # keep looping until we break out of the loop
    try:
        SelectedSheet = int(input("\nSelect Sheet ID: "))
        if (SelectedSheet -1) < len(MySheets):
            break    # exit the loop if the previous line succeeded
        else:
            print("Sheet ID is out of range, highest sheet ID is: " + str(len(MySheets)))
    except ValueError:
        print("Please enter an integer!")

# Set the workingsheet to the selected sheet
SelectSheetName = MySheets[SelectedSheet-1][1]
myworksheet = myworkbook[SelectSheetName]

# Dispaly the selected sheet
print("Selected Sheet Name: " + SelectSheetName)

### Obtain left column, starting row and the number of rows in order to calculate loop range
KeyColumn = input("Key Column (Column with IP Addresses):")
StartRow = input("Starting Row:")
NumRows = input("Number of Rows: ")
EndRowNum = int(StartRow) + int(NumRows)
ColIndexNum = letters.index(KeyColumn.lower())
DescColumn = letters[ColIndexNum + 1] # Increments row column (e.g. A + 1 = B)

# Loop To read values into a  2D array called MyIPs
for loop in range(int(StartRow), EndRowNum):
    my_ip_cell = KeyColumn + str(loop) #Sets the first cell to be read 
    my_desc_cell = DescColumn.upper() + str(loop) #Sets the second cell to be read
    MyIPs.append([myworksheet[my_ip_cell].value, myworksheet[my_desc_cell].value])

# Reset LoopCounter used previous in sheet indexing
LoopCounter = 0

# Loop to display a subset of the output for validation purposes, before proceeding with an upload
# First X and last X are printed to ensure the range looks correct
# LoopCouter is used to control how many lines are printed
for x in range(len(MyIPs)):
    LoopCounter += 1
    if LoopCounter <= PreviewAmount: # Print if outputted lines are less than the set preview amount
        print(str(MyIPs[x][0]) + "    ---- " + str(MyIPs[x][1]))
        if LoopCounter == PreviewAmount: print("\n ------- Ommited ------- \n") # Print dividing line
    elif x > (len(MyIPs)-PreviewAmount): # Print last X amount of entries
        print(str(MyIPs[x][0]) + "    ---- " + str(MyIPs[x][1])) 
####

